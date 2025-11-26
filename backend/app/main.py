from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import date
import calendar

# Importações internas do projeto
from .database import engine, get_db
from . import models, schemas, crud

# Cria as tabelas no Banco de Dados se não existirem
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Gerenciador de Escalas Multi-Lojas")

# --- CONFIGURAÇÃO DO CORS ---
# Permite que o Frontend (React) fale com o Backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# ROTAS DE GERENCIAMENTO DE ESCALAS (HOME)
# ==========================================

@app.get("/escalas/", response_model=List[schemas.Escala])
def listar_todas_escalas(db: Session = Depends(get_db)):
    """Retorna a lista de todas as escalas criadas para mostrar na Home."""
    return crud.list_escalas(db)

@app.post("/escalas/", response_model=schemas.Escala)
def criar_nova_escala(escala: schemas.EscalaCreate, db: Session = Depends(get_db)):
    """Cria uma nova escala. Pode ser vazia ou já vir com dados."""
    return crud.create_escala(db, escala)

@app.get("/escalas/{escala_id}", response_model=schemas.Escala)
def abrir_escala(escala_id: int, db: Session = Depends(get_db)):
    """Carrega os dados de uma escala específica."""
    escala = crud.get_escala(db, escala_id)
    if not escala:
        raise HTTPException(status_code=404, detail="Escala não encontrada")
    return escala

@app.put("/escalas/{escala_id}")
def salvar_escala_existente(escala_id: int, dados: schemas.EscalaCreate, db: Session = Depends(get_db)):
    """Salva as alterações (grade e cores) de uma escala."""
    return crud.update_escala(db, escala_id, dados)

@app.delete("/escalas/{escala_id}")
def excluir_escala(escala_id: int, db: Session = Depends(get_db)):
    """Deleta uma escala e todos os seus funcionários."""
    sucesso = crud.delete_escala(db, escala_id)
    if not sucesso:
        raise HTTPException(status_code=404, detail="Escala não encontrada")
    return {"ok": True}

@app.post("/escalas/{escala_id}/duplicar")
def dar_continuidade(escala_id: int, novo_mes: int, novo_ano: int, db: Session = Depends(get_db)):
    """
    Rota Mágica: Pega uma escala existente, copia os funcionários e cria uma nova
    no mês seguinte, mantendo o histórico intacto.
    """
    nova = crud.duplicar_escala(db, escala_id, novo_mes, novo_ano)
    if not nova:
        raise HTTPException(status_code=404, detail="Escala origem não encontrada")
    return nova

# ==========================================
# ROTAS DE FUNCIONÁRIOS
# ==========================================

@app.get("/escalas/{escala_id}/funcionarios", response_model=List[schemas.Funcionario])
def listar_funcionarios_da_escala(escala_id: int, db: Session = Depends(get_db)):
    """Lista apenas os funcionários que pertencem a esta escala específica."""
    return crud.get_funcionarios_por_escala(db, escala_id)

@app.post("/funcionarios/", response_model=schemas.Funcionario)
def criar_funcionario(funcionario: schemas.FuncionarioCreate, db: Session = Depends(get_db)):
    """Cadastra um funcionário vinculado a uma escala."""
    return crud.create_funcionario(db, funcionario)

@app.put("/funcionarios/{func_id}", response_model=schemas.Funcionario)
def atualizar_funcionario(func_id: int, funcionario: schemas.FuncionarioCreate, db: Session = Depends(get_db)):
    """Atualiza dados do funcionário (nome, cargo, equipe)."""
    return crud.update_funcionario(db, func_id, funcionario)

@app.delete("/funcionarios/{func_id}")
def deletar_funcionario(func_id: int, db: Session = Depends(get_db)):
    """Remove um funcionário."""
    crud.delete_funcionario(db, func_id)
    return {"ok": True}

# ==========================================
# EXPORTAÇÃO PARA EXCEL
# ==========================================

@app.get("/exportar_excel/{escala_id}")
def exportar_excel(escala_id: int, db: Session = Depends(get_db)):
    """
    Gera o arquivo XLSX baseado no ID da escala salva.
    Isso garante que baixamos exatamente o que está no banco, com os nomes corretos.
    """
    # 1. Busca dados do Banco
    escala = crud.get_escala(db, escala_id)
    if not escala:
        raise HTTPException(status_code=404, detail="Escala não encontrada")
    
    funcionarios = crud.get_funcionarios_por_escala(db, escala_id)
    
    # Mapeamentos para preencher o Excel
    mapa_nomes = {str(f.id): f.nome for f in funcionarios}
    mapa_cargos = {str(f.id): f.cargo for f in funcionarios}
    dados_grade = escala.dados_escala or {}
    cores_legenda = escala.legenda_cores or {}

    # 2. Configura o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{escala.nome}"[:30] # Limite de caracteres do Excel
    
    # Estilos
    fonte_negrito = Font(bold=True)
    fonte_branca = Font(bold=True, color="FFFFFF")
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    borda_fina = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Prepara cores (remove o # do Hex)
    cores_formatadas = {k: v.replace("#", "") for k, v in cores_legenda.items()}

    # Configuração de Calendário
    ultimo_dia = calendar.monthrange(escala.ano, escala.mes)[1]
    dias_semana = ["DOM", "SEG", "TER", "QUA", "QUI", "SEX", "SÁB"]

    # --- CABEÇALHOS ---
    ws.cell(row=1, column=1, value="Funcionário").font = fonte_negrito
    ws.cell(row=1, column=2, value="Cargo").font = fonte_negrito
    
    # Dias da Semana
    for dia in range(1, ultimo_dia + 1):
        col = 2 + dia
        data_atual = date(escala.ano, escala.mes, dia)
        idx_sem = (data_atual.isoweekday()) % 7 # Ajuste para bater com lista (0=Dom)
        
        celula = ws.cell(row=1, column=col, value=dias_semana[idx_sem])
        celula.alignment = alinhamento_centro
        celula.border = borda_fina
        
        if idx_sem == 0: # Domingo
            celula.fill = PatternFill(start_color="FFFF00", fill_type="solid")
            celula.font = Font(bold=True, color="000000")
        else:
            celula.fill = PatternFill(start_color="E0E0E0", fill_type="solid")

    # Dias do Mês (Números)
    for dia in range(1, ultimo_dia + 1):
        col = 2 + dia
        celula = ws.cell(row=2, column=col, value=dia)
        celula.alignment = alinhamento_centro
        celula.border = borda_fina

    # --- DADOS DOS FUNCIONÁRIOS ---
    linha_atual = 3
    
    # Itera sobre os funcionários REAIS da escala
    for func in funcionarios:
        func_id_str = str(func.id)
        
        # Coluna 1 e 2
        ws.cell(row=linha_atual, column=1, value=func.nome).border = borda_fina
        ws.cell(row=linha_atual, column=2, value=func.cargo).border = borda_fina
        
        # Dias
        dias_do_func = dados_grade.get(func_id_str, {})
        
        for dia in range(1, ultimo_dia + 1):
            col = 2 + dia
            status = dias_do_func.get(str(dia), "-")
            
            celula = ws.cell(row=linha_atual, column=col, value=status)
            celula.alignment = alinhamento_centro
            celula.border = borda_fina
            
            # Pintura
            if status in cores_formatadas and status != "-":
                cor_hex = cores_formatadas[status]
                try:
                    celula.fill = PatternFill(start_color=cor_hex, fill_type="solid")
                    # Texto Branco para M e R
                    if status in ["M", "R"]:
                        celula.font = fonte_branca
                except:
                    pass # Se a cor for inválida, ignora

        linha_atual += 1

    # Ajuste de Largura
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    for col in range(3, 35):
        letra = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letra].width = 4

    # 3. Retorna o Arquivo
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Escala_{escala.nome}_{escala.mes}_{escala.ano}.xlsx"
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )